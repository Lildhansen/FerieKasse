#imports
import random
from ast import Raise
import openpyxl
from openpyxl.styles import Font, Color 
import os

#own modules
import utilities.util as util
import utilities.constants as const
from classes.Team import Team
from classes.Player import Player


class Excel:
    def __init__(self,leagues=[]):
        self.leagues = leagues
        self.players = util.getPlayerObjectsFromFile()
        self.playersTeamsDict = {}
        self.leaguesAndColors = {}
        self.addColorsToLeagues(self.leagues)
    
    #decide the colors for each league (to be used in the excel sheet)
    def addColorsToLeagues(self,leagues):
        availableColors = const.AVAILABLE_COLORS.copy()
        random.shuffle(const.AVAILABLE_COLORS)
        for league in leagues:
            self.leaguesAndColors[league.name] = availableColors[0]
            availableColors.pop(0)
            
    #sets up the excel file. This is done when starting a game
    def setupExcelFile(self):
        wb = openpyxl.Workbook() # new workbook
        ws = wb.active #new worksheet
        ws.title = "Feriekasse"
        row = 1
        column = 1
        cellWidths = []
        for player in self.players:
            playerNameLength = len(player.name)
            cellWidth = playerNameLength if playerNameLength > len("Total:") else len("Total:")
            #first column set
            ws.cell(row,column,player.name)
            for team in player.teams:
                teamNameLength = len(team.name)
                if teamNameLength > cellWidth:
                    cellWidth = teamNameLength
                row += 1
                ws.cell(row,column,team.name)
                ws[f"{util.numberToExcelColumn(column)}{row}"].font = Font(color=Color(self.leaguesAndColors[team.leagueName]))
            cellWidths.append(cellWidth)
            row += 1
            ws.cell(row,column,"Total:")
            #second column set
            row = 1
            column += 1
            ws.cell(row,column,"Point:")
            for team in player.teams:
                row += 1
                pointLocation = util.numberToExcelColumn(column) + str(row)
                ws[pointLocation] = "=0"
            row += 1
            ws.cell(row,column,f"=SUM({util.numberToExcelColumn(column)}{row - len(player.teams)}:{util.numberToExcelColumn(column)}{row-1})")
            row = 1
            column += 1
        for league in self.leagues:
            row += 1
            ws.cell(row,column,league.name)
            ws[f"{util.numberToExcelColumn(column)}{row}"].font = Font(color=Color(self.leaguesAndColors[league.name]))
        
        #set widths:
        i = 1
        #ws.column_dimensions.width = 100 
        for cellWidth in cellWidths:
            ws.column_dimensions[util.numberToExcelColumn(i)].width = cellWidth * 1.23
            i += 2
        
        wb.save(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
        wb.close()
    
    #updates the excel file. this is done when a game is in progress
    def updateExcelFile(self,players):
        wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
        ws = wb.active #current worksheet
        column = 1
        for player in players:
            for match in player.matches:
                self.updateTeamPointsInColumn(match,ws,column)
            column += 2
        if const.FOUR_GOAL_WIN_RULE:
            self.addFourGoalWinBonusPoints(ws,players)
        wb.save(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
        wb.close()
    
    #adds bonus points for four goal win rule
    def addFourGoalWinBonusPoints(self,ws,players):
        column = 1
        row = 1
        for player in players:
            while True:
                if player.name == ws.cell(row,column).value:
                    break
                column += 2
            for team in player.teams:
                row = 1
                if team.bonusPoints == 0:
                    continue
                cell = ws.cell(row,column)
                if team.name == cell.value:
                    pointCell = ws.cell(row,column+1)
                    pointCell.value += "+" + str(team.bonusPoints)
                else:
                    row += 1   
        
    #deletes the excel file. this is done when a game is over
    def deleteExcelFile(self):
        if (os.path.isfile(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")):
            os.remove(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
            
    #updates the points in the excel file for a single match
    def updateTeamPointsInColumn(self,match,ws,column):
        row = 2
        while True:
            cell = ws.cell(row,column)
            if cell.value == "Total:":
                Raise(Exception("team not found in excel file"))
            if cell.value == match.homeTeam or cell.value == match.awayTeam:
                pointCell = ws.cell(row,column+1)
                pointCell.value += "+" + str(match.points)
                break
            row += 1
            
    #Gets the total points for a player and returns that player
    def getPlayerScoreFromExcelFile(self,player):
        wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
        ws = wb.active #current worksheet
        column = 1
        #finds the player in the excel file
        while True:
            if player.name == ws.cell(1,column).value:
                player = Player(player.name)
                break
            if ws.cell(1,column).value == "":
                Raise(Exception("player not found in excel file"))
            column += 2
        #finds the total points of the player - does this for all teams rather than looking at total, since the result of total is an equation (SUM[A1:A6])
        row = 2
        totalPoints = 0
        while True:
            if ws.cell(row,column).value == "Total:":
                break
            totalPoints += util.getSumOfExcelCell(ws.cell(row,column+1).value)
            row += 1
        player.totalPoints += totalPoints
        wb.close()
        return player
        
    #gets the team that has the most points
    def getHighestScoreTeam(self):
        def getHighestScoreTeamHelper(teams):
            teams.sort(key=lambda team: team.points, reverse=True)
            return teams[0]
            
        return self.getTeam(getHighestScoreTeamHelper)
        
    #gets the team that has the least points
    def getLowestScoreTeam(self):
        def getLowestScoreTeamHelper(teams):
            teams.sort(key=lambda team: team.points, reverse=False)
            return teams[0]
        return self.getTeam(getLowestScoreTeamHelper) 
        
    #gets a team based on a function that returns a specific team
    def getTeam(self,getSpecificTeamFunction):
        teams = []
        wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
        ws = wb.active #current worksheet
        column = 1
        while ws.cell(2,column+1).value != None:
            row = 2
            playerName = ws.cell(row-1,column).value
            while ws.cell(row+1,column).value != None:
                teamName = ws.cell(row,column).value
                newTeam = Team(teamName,playerName)
                newTeam.points = util.getSumOfExcelCell(ws.cell(row,column+1).value)
                teams.append(newTeam)
                row += 1
            column += 2
        wb.close()
        return getSpecificTeamFunction(teams)