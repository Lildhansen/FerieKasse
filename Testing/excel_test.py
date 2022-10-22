import pytest
from Excel import Excel
from classes.League import League
from classes.Match import Match
from classes.Player import Player
import utilities.constants as const
import os
import openpyxl

def setupMockExcelFile():
    const.FERIEKASSE_NAME = "unitTests/excelTest"
    if os.path.exists(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx"):
        os.remove(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
    league1 = League("league1","country1")
    league2 = League("league2","country2")
    league3 = League("league3","country3")
    return Excel([league1,league2,league3])

def changeValuesForGetXXXScoreTeamAndReturnExcel():
    excel = setupMockExcelFile()
    excel.setupExcelFile()
    match1 = Match()
    match1.points = 20
    match1.homeTeam = "team1"
    match2 = Match()
    match2.points = 10
    match2.homeTeam = "team2"
    match3 = Match()
    match3.points = 30
    match3.homeTeam = "team3"
    match4 = Match()
    match4.points = 40
    match4.homeTeam = "team4"
    wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
    ws = wb.active #current worksheet
    excel.updateTeamPointsInColumn(match1,ws,1,False)
    excel.updateTeamPointsInColumn(match3,ws,1,False)
    excel.updateTeamPointsInColumn(match2,ws,3,False)
    excel.updateTeamPointsInColumn(match4,ws,3,False)
    wb.save(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
    wb.close()
    return excel

def test_excel_init_adds_colors_to_leagues():
    excel = setupMockExcelFile()
    assert excel.leaguesAndColors["league1"] != None
    assert excel.leaguesAndColors["league2"] != None
    assert excel.leaguesAndColors["league3"] != None

def test_excel_file_can_be_deleted():
    excel = setupMockExcelFile()
    with open(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx", "w"):
        pass
    assert os.path.exists(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
    excel.deleteExcelFile()
    assert not os.path.exists(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
 
#integration tests down here:
    
def test_setupExcelFile_creates_excel_file():
    excel = setupMockExcelFile()
    excel.setupExcelFile()
    assert os.path.exists(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")

#this is a beautiful test
def test_setupExcelFile_formats_excel_file_correctly():
    excel = setupMockExcelFile()
    excel.setupExcelFile()
    
    #load newly created excel file
    wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
    ws = wb.active #current worksheet
    assert ws.cell(1,1).value == excel.players[0].name
    assert ws.cell(2,1).value == excel.players[0].teams[0].name
    assert ws.cell(3,1).value == excel.players[0].teams[1].name
    assert ws.cell(4,1).value == "Total:"
    
    assert ws.cell(1,2).value == "Point:"
    assert ws.cell(2,2).value == "=0"
    assert ws.cell(3,2).value == "=0"
    assert ws.cell(4,2).value == "=SUM(B2:B3)"
    
    assert ws.cell(1,3).value == excel.players[1].name
    assert ws.cell(2,3).value == excel.players[1].teams[0].name
    assert ws.cell(3,3).value == excel.players[1].teams[1].name
    assert ws.cell(4,3).value == "Total:"
    
    assert ws.cell(1,4).value == "Point:"
    assert ws.cell(2,4).value == "=0"
    assert ws.cell(3,4).value == "=0"
    assert ws.cell(4,4).value == "=SUM(D2:D3)"
    
    assert ws.cell(1,5).value == None
    assert ws.cell(2,5).value == excel.leagues[0].name
    assert ws.cell(3,5).value == excel.leagues[1].name
    assert ws.cell(4,5).value == excel.leagues[2].name
    wb.close()
    
def test_updateExcelFile_updates_excel_file_correctly():
    excel = setupMockExcelFile()
    excel.setupExcelFile()
    player1 = Player("player1")
    match1 = Match()
    match1.points = 10
    match1.homeTeam = "team1"
    match3 = Match()
    match3.points = 30
    match3.homeTeam = "team3"
    player1.matches = [match1,match3]
    player2 = Player("player2")
    match2 = Match()
    match2.points = 20
    match2.homeTeam = "team2"
    match4 = Match()
    match4.points = 40
    match4.homeTeam = "team4"
    player2.matches = [match2,match4]
    players = [player1,player2]
    excel.updateExcelFile(players)
    wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
    ws = wb.active #current worksheet
    assert ws.cell(2,2).value == "=0+10"
    assert ws.cell(3,2).value == "=0+30"
    assert ws.cell(2,4).value == "=0+20"
    assert ws.cell(3,4).value == "=0+40"
    wb.close()

def test_updateTeamPointsInColumn_updates_correct_cell_in_column_with_correct_value():
    excel = setupMockExcelFile()
    excel.setupExcelFile()
    match = Match()
    match.points = 10
    match.homeTeam = "team1"
    wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
    ws = wb.active #current worksheet
    excel.updateTeamPointsInColumn(match,ws,1,False)
    assert ws.cell(2,2).value == "=0+10"
    wb.close() 

def test_setPlayerTotalPointsVariableAndGetPlayerFromExcelFile_sets_correct_player_total_points_and_gets_player():
    excel = setupMockExcelFile()
    excel.setupExcelFile()
    match1 = Match()
    match1.points = 20
    match1.homeTeam = "team1"
    match2 = Match()
    match2.points = 30
    match2.homeTeam = "team3"
    player = Player("player1")
    wb = openpyxl.load_workbook(fr'data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx')
    ws = wb.active #current worksheet
    excel.updateTeamPointsInColumn(match1,ws,1,False)
    excel.updateTeamPointsInColumn(match2,ws,1,False)
    wb.save(fr"data/{const.FERIEKASSE_NAME}/Feriekasse.xlsx")
    wb.close()
    returnedPlayer = excel.setPlayerTotalPointsVariableAndGetPlayerFromExcelFile(player)
    assert returnedPlayer.totalPoints == 50
    assert player == returnedPlayer

def test_getLowestScoreTeam_return_team_with_lowest_score():
    excel = changeValuesForGetXXXScoreTeamAndReturnExcel()
    team = excel.getLowestScoreTeam()
    assert team.name == "team2"
    
    
def test_getHighestScoreTeam_return_team_with_highest_score():
    excel = changeValuesForGetXXXScoreTeamAndReturnExcel()
    team = excel.getHighestScoreTeam()
    assert team.name == "team4"










#const.FERIEKASSE_NAME = "unitTests/extraBodyTrailingTrue"